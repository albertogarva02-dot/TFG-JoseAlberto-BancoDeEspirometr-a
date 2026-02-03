#_____________________________________________________________________________________________________________________
#    Nombre del Script: exportar_excel.py
#    Autor: José Alberto García Valiño
#    Fecha: 07-12-2025
#    Descripción:
#        Módulo para la generación de informes técnicos.
#        Utiliza Pandas y OpenPyXL para crear hojas de cálculo,
#        incluyen tablas de datos de pacientes, métricas calculadas y gráficos de las curvas resultantes.
#______________________________________________________________________________________________________________________

__author__ = "José Alberto García Valiño"
__copyright__ = "Copyright 2025, José Alberto García Valiño"
__credits__ = ["José Alberto García Valiño"]
__license__ = "MIT"
__version__ = "1.0"
__status__ = "Beta"

import pandas as pd
import ast
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

TITULO_FUENTE = Font(name='Calibri', size=18, bold=True, color="FFFFFF")
TITULO_RELLENO = PatternFill(start_color="004C99", end_color="004C99", fill_type="solid")
ENCABEZADO_FUENTE = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
ENCABEZADO_RELLENO = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
TABLA_ESTILO = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
GRAFICA_COLOR_LINEA = "0078D4"

def _auto_ajustar_columnas(hoja_calculo):
    for idx_col in range(1, hoja_calculo.max_column + 1):
        letra_columna = get_column_letter(idx_col)
        longitud_maxima = 0
        for celda in hoja_calculo[letra_columna]:
            try:
                if len(str(celda.value)) > longitud_maxima:
                    longitud_maxima = len(str(celda.value))
            except:
                pass
        ancho_ajustado = (longitud_maxima + 2)
        hoja_calculo.column_dimensions[letra_columna].width = ancho_ajustado

def _crear_hoja_informe(escritor, datos):
    #límite de 31 caracteres para nombres de hojas (excel)
    nombre_sim = datos[0]
    nombre_hoja = f"Informe {nombre_sim}"[:31]

    #PREPARAR DE DATOS
    df_info = pd.DataFrame({
        "Parámetro": ["Nombre Simulación", "Nombre Paciente", "Edad [años]", "Peso [Kg]", "Altura [cm]", "Sexo", "DNI", "Fumador", "Fecha Creación", "Patología"],
        "Valor": [datos[0], datos[1], datos[2], datos[3], datos[8], datos[4], datos[5], datos[6], datos[7], datos[9]]
    })
    valores_lista = ast.literal_eval(datos[10])
    df_valores = pd.DataFrame(valores_lista, columns=['Volumen (L)', 'Flujo (L/s)'])
    
    #CALCULO METRICAS CLAVE (añadir mas/o crer en base de datos nuevas columas con los valores ya calculados)
    fvc = df_valores['Volumen (L)'].max() if not df_valores.empty else 0
    pef = df_valores['Flujo (L/s)'].max() if not df_valores.empty else 0
    df_metricas = pd.DataFrame({
        "Métrica Clave": ["Capacidad Vital Forzada (FVC)", "Flujo Espiratorio Pico (PEF)"],
        "Resultado": [f"{fvc:.2f} L", f"{pef:.2f} L/s"]
    })

    #ESCRITURA EN EXCEL
    df_info.to_excel(escritor, sheet_name=nombre_hoja, index=False, startrow=2, startcol=0)
    df_metricas.to_excel(escritor, sheet_name=nombre_hoja, index=False, startrow=2, startcol=4)
    df_valores.to_excel(escritor, sheet_name=nombre_hoja, index=False, startrow=len(df_info) + 5, startcol=0)

    libro_trabajo = escritor.book
    hoja_calculo = escritor.sheets[nombre_hoja]
    hoja_calculo.sheet_view.showGridLines = False

    #titulo principal
    hoja_calculo.merge_cells('A1:G1')
    celda_titulo = hoja_calculo['A1']
    celda_titulo.value = f"Informe de Simulación: {nombre_sim}"
    celda_titulo.font = TITULO_FUENTE
    celda_titulo.fill = TITULO_RELLENO
    celda_titulo.alignment = Alignment(horizontal='center', vertical='center')

    #cabeceras
    for ref_cabecera in ['A3', 'B3', 'E3', 'F3']:
        celda = hoja_calculo[ref_cabecera]
        celda.font = ENCABEZADO_FUENTE
        celda.fill = ENCABEZADO_RELLENO

    #tablas
    tabla_info = Table(displayName=f"Info_{nombre_sim.replace(' ','_')}", ref=f"A3:B{len(df_info)+3}")
    tabla_info.tableStyleInfo = TABLA_ESTILO
    hoja_calculo.add_table(tabla_info)

    tabla_metricas = Table(displayName=f"Metricas_{nombre_sim.replace(' ','_')}", ref=f"E3:F{len(df_metricas)+3}")
    tabla_metricas.tableStyleInfo = TABLA_ESTILO
    hoja_calculo.add_table(tabla_metricas)

    ref_tabla_valores = f"A{len(df_info) + 6}:B{len(df_info) + 6 + len(df_valores)}"
    tabla_valores = Table(displayName=f"Valores_{nombre_sim.replace(' ','_')}", ref=ref_tabla_valores)
    tabla_valores.tableStyleInfo = TABLA_ESTILO
    hoja_calculo.add_table(tabla_valores)

    #grafica
    grafica = ScatterChart()
    grafica.title = "Bucle Flujo-Volumen"
    grafica.style = 2
    grafica.x_axis.title = "Volumen (L)"
    grafica.y_axis.title = "Flujo (L/s)"
    grafica.legend = None

    valores_x = Reference(hoja_calculo, min_col=1, min_row=len(df_info) + 7, max_row=len(df_info) + 6 + len(df_valores))
    valores_y = Reference(hoja_calculo, min_col=2, min_row=len(df_info) + 7, max_row=len(df_info) + 6 + len(df_valores))

    serie = Series(valores_y, valores_x, title_from_data=False)
    serie.graphicalProperties.line.solidFill = GRAFICA_COLOR_LINEA
    serie.graphicalProperties.line.width = 25000
    serie.marker.symbol = "none"
    grafica.series.append(serie)
    hoja_calculo.add_chart(grafica, "D8")

    #ajuste columnas
    _auto_ajustar_columnas(hoja_calculo)


def exportar_a_excel(datos, ruta_archivo):#crea archivo y añade hoja
    try:
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as escritor:
            _crear_hoja_informe(escritor, datos)
        return True, f"Informe profesional guardado en: {ruta_archivo}"
    
    except Exception as e:
        print(f"ERROR DETALLADO AL CREAR EXCEL (único): {e}")
        return False, f"Error al generar el informe Excel: {e}"


def exportar_comparacion_excel(lista_de_datos_completos, ruta_archivo):#crear hoja conjunta y una hoja por cada simu
    try:
        with pd.ExcelWriter(ruta_archivo, engine='openpyxl') as escritor:
            
            #HOJA DE RESUMEN
            nombre_hoja_resumen = "Comparativa"
            datos_resumen = []
            
            #datos para la tabla de resumen
            for datos in lista_de_datos_completos:
                nombre_sim = datos[0]
                patologia = datos[9]
                edad = datos[2]
                sexo = datos[4]
                
                valores_lista = ast.literal_eval(datos[10])
                df_temp = pd.DataFrame(valores_lista, columns=['Volumen (L)', 'Flujo (L/s)'])
                fvc = df_temp['Volumen (L)'].max() if not df_temp.empty else 0
                pef = df_temp['Flujo (L/s)'].max() if not df_temp.empty else 0
                
                datos_resumen.append([nombre_sim, f"{fvc:.2f}", f"{pef:.2f}", patologia, edad, sexo])
            
            df_resumen = pd.DataFrame(datos_resumen, columns=["Simulación", "FVC (L)", "PEF (L/s)", "Patología", "Edad", "Sexo"])
            df_resumen.to_excel(escritor, sheet_name=nombre_hoja_resumen, index=False, startrow=2, startcol=0)
            
            #ESTILOS DE HOJA DE RESUMEN
            libro_trabajo = escritor.book
            hoja_resumen = escritor.sheets[nombre_hoja_resumen]
            hoja_resumen.sheet_view.showGridLines = False

            hoja_resumen.merge_cells('A1:G1')
            celda_titulo = hoja_resumen['A1']
            celda_titulo.value = "Informe Comparativo de Simulaciones"
            celda_titulo.font = TITULO_FUENTE
            celda_titulo.fill = TITULO_RELLENO
            celda_titulo.alignment = Alignment(horizontal='center', vertical='center')

            for num_col in range(df_resumen.shape[1]):
                celda = hoja_resumen.cell(row=3, column=num_col + 1)
                celda.font = ENCABEZADO_FUENTE
                celda.fill = ENCABEZADO_RELLENO

            tabla_resumen = Table(displayName="ResumenComparativa", ref=f"A3:{get_column_letter(df_resumen.shape[1])}{len(df_resumen)+3}")
            tabla_resumen.tableStyleInfo = TABLA_ESTILO
            hoja_resumen.add_table(tabla_resumen)

            #GRAFICA COMPARATIVA
            grafica_comp = ScatterChart()
            grafica_comp.title = "Comparativa Flujo-Volumen"
            grafica_comp.x_axis.title = "Volumen (L)"
            grafica_comp.y_axis.title = "Flujo (L/s)"
            grafica_comp.legend.position = "r" 

            fila_actual = len(df_resumen) + 6
            hoja_resumen[f"A{fila_actual-1}"] = "Datos Crudos para Gráfica:"
            
            desplazamiento_col = 0
            for i, datos in enumerate(lista_de_datos_completos):
                nombre_sim = datos[0]
                valores_lista = ast.literal_eval(datos[10])
                df_valores = pd.DataFrame(valores_lista, columns=['Volumen (L)', 'Flujo (L/s)'])
                
                #cabeceras para estos datos
                hoja_resumen.cell(row=fila_actual, column=desplazamiento_col + 1, value=f"{nombre_sim}-Vol")
                hoja_resumen.cell(row=fila_actual, column=desplazamiento_col + 2, value=f"{nombre_sim}-Flujo")
                
                #escribir datos
                for idx_fila, datos_fila in enumerate(df_valores.values):
                    hoja_resumen.cell(row=fila_actual + 1 + idx_fila, column=desplazamiento_col + 1, value=datos_fila[0])
                    hoja_resumen.cell(row=fila_actual + 1 + idx_fila, column=desplazamiento_col + 2, value=datos_fila[1])
                
                #referencias
                vals_x = Reference(hoja_resumen, min_col=desplazamiento_col + 1, min_row=fila_actual + 1, max_row=fila_actual + len(df_valores))
                vals_y = Reference(hoja_resumen, min_col=desplazamiento_col + 2, min_row=fila_actual + 1, max_row=fila_actual + len(df_valores))
                
                serie = Series(vals_y, vals_x, title=nombre_sim)
                serie.graphicalProperties.line.width = 20000
                serie.marker.symbol = "none"
                grafica_comp.series.append(serie)
                
                desplazamiento_col += 3 #siguiente par de columnas

            hoja_resumen.add_chart(grafica_comp, f"{get_column_letter(df_resumen.shape[1] + 2)}3")
            _auto_ajustar_columnas(hoja_resumen)

            #CREAR HOJAS INDIVIDUALES
            for datos in lista_de_datos_completos:
                _crear_hoja_informe(escritor, datos)
        
        return True, f"Informe comparativo guardado en: {ruta_archivo}"
    
    except Exception as e:
        return False, f"Error al generar el informe comparativo: {e}"